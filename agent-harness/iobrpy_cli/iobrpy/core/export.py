"""
export.py - Export functionality for IOBRpy CLI harness.

This module provides utilities for exporting IOBRpy results in various formats,
including JSON, Excel, and summary reports.
"""

import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List, Any, Union

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class ExportResult:
    """Result of an export operation."""

    success: bool
    output_path: Optional[str] = None
    message: str = ""
    file_count: int = 0


class Exporter:
    """Handles exporting IOBRpy results."""

    def __init__(self, output_dir: Path):
        """
        Initialize exporter.

        Args:
            output_dir: Base output directory
        """
        self.output_dir = Path(output_dir)

    def export_json(
        self,
        data: Union[Dict, List],
        output_path: Optional[Path] = None,
        pretty: bool = True,
    ) -> ExportResult:
        """
        Export data to JSON format.

        Args:
            data: Data to export (dict or list)
            output_path: Output file path (default: output_dir/export.json)
            pretty: Pretty-print JSON

        Returns:
            ExportResult
        """
        if output_path is None:
            output_path = self.output_dir / f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        else:
            output_path = Path(output_path)

        try:
            with open(output_path, 'w') as f:
                json.dump(data, f, indent=2 if pretty else None)
            return ExportResult(
                success=True,
                output_path=str(output_path),
                message=f"Successfully exported to {output_path}",
                file_count=1,
            )
        except Exception as e:
            return ExportResult(
                success=False,
                message=f"Failed to export JSON: {str(e)}",
            )

    def export_dataframe(
        self,
        df: 'pd.DataFrame',
        output_path: Optional[Path] = None,
        format: str = 'csv',
        **kwargs,
    ) -> ExportResult:
        """
        Export pandas DataFrame.

        Args:
            df: DataFrame to export
            output_path: Output file path
            format: Output format ('csv', 'tsv', 'excel', 'json')
            **kwargs: Additional arguments for pandas export function

        Returns:
            ExportResult
        """
        if not HAS_PANDAS:
            return ExportResult(
                success=False,
                message="pandas is not available",
            )

        if output_path is None:
            output_path = self.output_dir / f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{format}"
        else:
            output_path = Path(output_path)

        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            if format == 'csv':
                df.to_csv(output_path, **kwargs)
            elif format == 'tsv':
                df.to_csv(output_path, sep='\t', **kwargs)
            elif format == 'excel':
                df.to_excel(output_path, **kwargs)
            elif format == 'json':
                df.to_json(output_path, **kwargs)
            else:
                return ExportResult(
                    success=False,
                    message=f"Unsupported format: {format}",
                )

            return ExportResult(
                success=True,
                output_path=str(output_path),
                message=f"Successfully exported to {output_path}",
                file_count=1,
            )
        except Exception as e:
            return ExportResult(
                success=False,
                message=f"Failed to export DataFrame: {str(e)}",
            )

    def export_tme_results(
        self,
        tme_dir: Path,
        output_path: Optional[Path] = None,
        format: str = 'excel',
    ) -> ExportResult:
        """
        Export TME deconvolution results.

        Args:
            tme_dir: Directory containing TME result files
            output_path: Output file path
            format: Output format ('excel', 'csv')

        Returns:
            ExportResult
        """
        if not HAS_PANDAS:
            return ExportResult(
                success=False,
                message="pandas is not available",
            )

        if format == 'excel' and not HAS_OPENPYXL:
            return ExportResult(
                success=False,
                message="openpyxl is not available for Excel export",
            )

        tme_dir = Path(tme_dir)
        if not tme_dir.exists():
            return ExportResult(
                success=False,
                message=f"TME directory not found: {tme_dir}",
            )

        # Find result files
        result_files = []
        for pattern in ['*_results.csv', '*_results.tsv', 'deconvo_merged.csv']:
            result_files.extend(tme_dir.glob(pattern))

        if not result_files:
            return ExportResult(
                success=False,
                message=f"No result files found in {tme_dir}",
            )

        if output_path is None:
            output_path = self.output_dir / f"tme_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{format}"
        else:
            output_path = Path(output_path)

        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            if format == 'excel':
                wb = Workbook()
                wb.remove(wb.active)

                for result_file in result_files:
                    df = pd.read_csv(result_file)
                    sheet_name = result_file.stem[:31]  # Excel sheet name limit
                    ws = wb.create_sheet(title=sheet_name)
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            ws.cell(row=r_idx, column=c_idx, value=value)

                wb.save(output_path)
            elif format in ('csv', 'tsv'):
                # For CSV/TSV, merge all files into one
                dfs = []
                for result_file in result_files:
                    df = pd.read_csv(result_file)
                    # Add method name as prefix to columns (except ID)
                    method = result_file.stem.replace('_results', '')
                    if 'ID' in df.columns:
                        id_col = df['ID']
                        df = df.drop(columns=['ID'])
                        df.columns = [f"{method}_{c}" for c in df.columns]
                        df['ID'] = id_col
                    dfs.append(df)

                merged = dfs[0]
                for df in dfs[1:]:
                    overlap = (set(merged.columns) & set(df.columns)) - {'ID'}
                    if overlap:
                        df = df.rename(columns={c: f"{c}_dup" for c in overlap})
                    merged = pd.merge(merged, df, on='ID', how='outer')

                merged.to_csv(output_path, sep='\t' if format == 'tsv' else ',', index=False)
            else:
                return ExportResult(
                    success=False,
                    message=f"Unsupported format: {format}",
                )

            return ExportResult(
                success=True,
                output_path=str(output_path),
                message=f"Successfully exported {len(result_files)} result files to {output_path}",
                file_count=len(result_files),
            )
        except Exception as e:
            return ExportResult(
                success=False,
                message=f"Failed to export TME results: {str(e)}",
            )

    def export_signature_scores(
        self,
        sig_file: Path,
        output_path: Optional[Path] = None,
        format: str = 'csv',
        top_n: Optional[int] = None,
    ) -> ExportResult:
        """
        Export signature scores with optional filtering.

        Args:
            sig_file: Signature scores file
            output_path: Output file path
            format: Output format ('csv', 'tsv', 'excel')
            top_n: Export only top N signatures by variance

        Returns:
            ExportResult
        """
        if not HAS_PANDAS:
            return ExportResult(
                success=False,
                message="pandas is not available",
            )

        sig_file = Path(sig_file)
        if not sig_file.exists():
            return ExportResult(
                success=False,
                message=f"Signature file not found: {sig_file}",
            )

        try:
            df = pd.read_csv(sig_file)

            if top_n is not None:
                # Calculate variance and keep top N
                if 'ID' in df.columns:
                    df_var = df.drop(columns=['ID']).var().nlargest(top_n)
                    selected_cols = ['ID'] + df_var.index.tolist()
                    df = df[selected_cols]
                else:
                    df_var = df.var().nlargest(top_n)
                    df = df[df_var.index]

            result = self.export_dataframe(df, output_path, format)
            if top_n:
                result.message = f"Exported top {top_n} signatures to {output_path}"
            return result

        except Exception as e:
            return ExportResult(
                success=False,
                message=f"Failed to export signature scores: {str(e)}",
            )

    def export_lr_scores(
        self,
        lr_file: Path,
        output_path: Optional[Path] = None,
        format: str = 'csv',
        threshold: Optional[float] = None,
    ) -> ExportResult:
        """
        Export ligand-receptor scores with optional filtering.

        Args:
            lr_file: LR scores file
            output_path: Output file path
            format: Output format ('csv', 'tsv', 'excel')
            threshold: Minimum score threshold

        Returns:
            ExportResult
        """
        if not HAS_PANDAS:
            return ExportResult(
                success=False,
                message="pandas is not available",
            )

        lr_file = Path(lr_file)
        if not lr_file.exists():
            return ExportResult(
                success=False,
                message=f"LR file not found: {lr_file}",
            )

        try:
            df = pd.read_csv(lr_file)

            if threshold is not None:
                # Filter by threshold
                # Assuming first column is LR pair name
                value_cols = df.columns[1:]
                df = df[(df[value_cols] >= threshold).any(axis=1)]

            result = self.export_dataframe(df, output_path, format)
            if threshold:
                result.message = f"Exported {len(df)} LR pairs (score >= {threshold}) to {output_path}"
            return result

        except Exception as e:
            return ExportResult(
                success=False,
                message=f"Failed to export LR scores: {str(e)}",
            )

    def create_summary_report(
        self,
        results_dir: Path,
        output_path: Optional[Path] = None,
    ) -> ExportResult:
        """
        Create a summary report of analysis results.

        Args:
            results_dir: Directory containing analysis results
            output_path: Output file path

        Returns:
            ExportResult
        """
        results_dir = Path(results_dir)
        if not results_dir.exists():
            return ExportResult(
                success=False,
                message=f"Results directory not found: {results_dir}",
            )

        if output_path is None:
            output_path = self.output_dir / f"summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        else:
            output_path = Path(output_path)

        try:
            summary = {
                'generated_at': datetime.now().isoformat(),
                'results_directory': str(results_dir),
                'sections': {},
            }

            # Scan for standard output directories
            for section in ['01-qc', '02-salmon', '02-star', '03-tpm',
                           '04-signatures', '05-tme', '06-LR_cal', '07-TCRBCR',
                           '01-signatures', '02-tme', '03-LR_cal']:
                section_path = results_dir / section
                if section_path.exists():
                    files = []
                    for f in section_path.iterdir():
                        if f.is_file():
                            files.append({
                                'name': f.name,
                                'size': f.stat().st_size,
                                'modified': datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
                            })
                    summary['sections'][section] = {
                        'path': str(section_path),
                        'file_count': len(files),
                        'files': files,
                    }

            # Try to read and summarize key results
            if HAS_PANDAS:
                # Summary TME results
                tme_paths = [
                    results_dir / '05-tme' / 'deconvo_merged.csv',
                    results_dir / '02-tme' / 'deconvo_merged.csv',
                ]
                for tme_path in tme_paths:
                    if tme_path.exists():
                        df = pd.read_csv(tme_path)
                        summary['tme_summary'] = {
                            'path': str(tme_path),
                            'sample_count': len(df),
                            'features': list(df.columns),
                        }
                        break

                # Summary signature scores
                sig_paths = [
                    results_dir / '04-signatures' / 'calculate_sig_score.csv',
                    results_dir / '01-signatures' / 'calculate_sig_score.csv',
                ]
                for sig_path in sig_paths:
                    if sig_path.exists():
                        df = pd.read_csv(sig_path)
                        summary['signature_summary'] = {
                            'path': str(sig_path),
                            'sample_count': len(df),
                            'signature_count': len(df.columns) - 1 if 'ID' in df.columns else len(df.columns),
                        }
                        break

            with open(output_path, 'w') as f:
                json.dump(summary, f, indent=2)

            return ExportResult(
                success=True,
                output_path=str(output_path),
                message=f"Summary report created at {output_path}",
                file_count=1,
            )

        except Exception as e:
            return ExportResult(
                success=False,
                message=f"Failed to create summary report: {str(e)}",
            )

    def export_directory(
        self,
        source_dir: Path,
        output_path: Optional[Path] = None,
        format: str = 'zip',
    ) -> ExportResult:
        """
        Export an entire directory.

        Args:
            source_dir: Source directory to export
            output_path: Output file path
            format: Archive format ('zip', 'tar', 'tar.gz')

        Returns:
            ExportResult
        """
        import shutil

        source_dir = Path(source_dir)
        if not source_dir.exists() or not source_dir.is_dir():
            return ExportResult(
                success=False,
                message=f"Source directory not found: {source_dir}",
            )

        if output_path is None:
            output_path = self.output_dir / f"{source_dir.name}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{format}"
        else:
            output_path = Path(output_path)

        try:
            if format == 'zip':
                shutil.make_archive(str(output_path.with_suffix('')), 'zip', source_dir)
            elif format == 'tar':
                shutil.make_archive(str(output_path.with_suffix('')), 'tar', source_dir)
            elif format in ('tar.gz', 'tgz'):
                shutil.make_archive(str(output_path.with_suffix('')), 'gztar', source_dir)
            else:
                return ExportResult(
                    success=False,
                    message=f"Unsupported format: {format}",
                )

            # Count files
            file_count = sum(1 for _ in source_dir.rglob('*') if _.is_file())

            return ExportResult(
                success=True,
                output_path=str(output_path),
                message=f"Successfully exported {file_count} files from {source_dir}",
                file_count=file_count,
            )

        except Exception as e:
            return ExportResult(
                success=False,
                message=f"Failed to export directory: {str(e)}",
            )
